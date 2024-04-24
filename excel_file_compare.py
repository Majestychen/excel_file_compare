from tkinter import *
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
import tkinter.messagebox

font_false = Font(size=9, bold=True, color="FF0000")  #bold：黑体
red_fill = PatternFill("solid", fgColor="C1CDCD") #PatternFill：图案填充；solid：固体；fgColor：前景色


def main():

    def fun_finish(shift, currentrow):
        data = ""
        if len(shift) > 0:
            for i in shift:
                data += str(i) + "."
        currentrow.fill = red_fill
        currentrow.font = font_false
        currentrow.value = data

    def fun_diff(start, end):
        arrShirt2 = []
        for i in start:
            a = 0
            for j in end:
                if i == j:
                    a += 1
            if a == 0:
                arrShirt2.append(i)
        return arrShirt2

    def selectExcelfile1():
        text1.delete(0, END)

        sfname1 = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx'), ('All Files', '*')])
        text1.insert(INSERT, sfname1)


    def selectExcelfile2():
        
        text2.delete(0, END)
        
        sfname2 = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx'), ('All Files', '*')])
        text2.insert(INSERT, sfname2)
    
    def doProcess():
        startFile1 = text1.get()
        startFile2 = text2.get()
        endFile = text3.get()
        wb1 = openpyxl.load_workbook(startFile1)
        wb2 = openpyxl.load_workbook(startFile2)
        # get workbook every son
#         sheet1 = wb['sheet1']
#         sheet2 = wb['sheet2']
        book1 = wb1.active
        book2 = wb2.active
        for rows in range(7, book2.max_row - 4):
            guige2 = book2['C' + str(rows)].value
            ishave = False
            for anorows in range(7, book1.max_row - 4):
                guige1 = book1['C' + str(anorows)].value
                if guige2 == guige1:
                    ishave = True
            if not ishave:
                book2['C' + str(rows)].fill = red_fill
                book2['C' + str(rows)].font = font_false
                book2['F' + str(rows)].fill = red_fill
                book2['F' + str(rows)].font = font_false
                book2['G' + str(rows)].fill = red_fill
                book2['G' + str(rows)].font = font_false

        for row in range(7, book1.max_row - 4):
            # 先判断sheet1 C列的子件规格的每一个和 sheet2中的 C列的子件规格进行对比
            guige1 = book1['C' + str(row)].value
            ishave = False
            currentAnoRow = -1
            for anorow in range(7, book2.max_row - 4):
                guige2 = book2['C' + str(anorow)].value
                if guige1 == guige2:
                    ishave = True
                    currentAnoRow = anorow
            if ishave:
                # 对比F/G的差异
                tp1 = book1['F' + str(row)].value
                tp2 = book2['F' + str(currentAnoRow)].value
                bm1 = book1['G' + str(row)].value
                bm2 = book2['G' + str(currentAnoRow)].value
                if tp1 is None or tp2 is None:
                    print('loading')
                else:
                    if tp1 != tp2:
                        print(type(tp1))
                        top1 = tp1.split(".")
                        top2 = tp2.split(".")
                        topshift1 = fun_diff(top1, top2)
                        topshift2 = fun_diff(top2, top1)
                        fun_finish(topshift1, book1['F' + str(row)])
                        fun_finish(topshift2, book2['F' + str(currentAnoRow)])
                if bm1 is None or bm2 is None:
                      print('loadnig again')
                else:
                    if bm1 != bm2:
                        print("bm1 is: ", bm1)
                        bottom1 = bm1.split(".")
                        bottom2 = bm2.split(".")
                        bottomshift1 = fun_diff(bottom1, bottom2)
                        bottomshift2 = fun_diff(bottom2, bottom1)
                        fun_finish(bottomshift1, book1['G' + str(row)])
                        fun_finish(bottomshift2, book2['G' + str(currentAnoRow)])
            else:
                book1['C' + str(row)].fill = red_fill
                book1['C' + str(row)].font = font_false
                book1['F' + str(row)].fill = red_fill
                book1['F' + str(row)].font = font_false
                book1['G' + str(row)].fill = red_fill
                book1['G' + str(row)].font = font_false
        else:
            tkinter.messagebox.showinfo('提示', '已处理完成,可在已选择的文件位置进行查看。')
        wb.save(endFile)

    def closeThisWindow():
        root.destroy()

    def saveExcelfile():
        text3.delete(0, END)
        sfname = filedialog.asksaveasfilename(title='选择保存的文件位置', filetype=[('Excel', '*.xlsx')])
        sfname = sfname + ".xlsx"
        text3.insert(INSERT, sfname)

    root = Tk()
    # 设置窗体标题
    root.title('文件比对器')
    # 设置窗口大小和位置
    root.geometry('500x300+570+200')
    label1 = Label(root, text='请选择要比对的文件:')
    text1 = Entry(root, bg='white', width=40)
    button1 = Button(root, text='浏览', width=4, height=1, command=selectExcelfile1)
    label2 = Label(root, text='请选择要比对的文件:')
    text2 = Entry(root, bg='white', width=40)
    button2 = Button(root, text='浏览', width=4, height=1, command=selectExcelfile2)
    text2 = Entry(root, bg='white', width=40)
    
    label3 = Label(root, text='请选择要保存的位置:')
    text3 = Entry(root, bg='white', width=40)
    button3 = Button(root, text='选择', width=4, height=1, command=saveExcelfile)
    button4 = Button(root, text='处理', width=8, command=doProcess)
    button5 = Button(root, text='退出', width=8, command=closeThisWindow)

    label1.pack()
    text1.pack()
    label2.pack()
    text2.pack()
    button1.pack()
    label3.pack()
    button3.pack()
    button3.pack()
    button4.pack()

    label1.place(x=5, y=30)
    text1.place(x=120, y=30)
    label2.place(x=5, y=60)
    text2.place(x=120, y=60)
    button1.place(x=400, y=26)
    button2.place(x=400, y=60)
    label3.place(x=5, y=100)
    text3.place(x=120, y=100)
    
    button3.place(x=400, y=100)
    button4.place(x=80, y=150)
    button5.place(x=360, y=150)
    root.mainloop()


if __name__ == "__main__":
    main()

